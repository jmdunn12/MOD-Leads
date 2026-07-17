#!/usr/bin/env python3
"""
================================================================================
LEAD ASSIGNMENT ENGINE  -  DMV home-improvement sales (AM / PM / EVE)
================================================================================
Assigns DC-Metro sales leads to reps for one shift using volume-weighted
performance scoring, opportunity tiers, drive time as a light factor, and
prior-slot drive chaining.

QUICK START (each shift)
------------------------
1. Gather (re-export fresh each day):
     - This slot's lead sheet      (SAP "Resource Planning" export .xlsx)
     - Power ranking               (2-week per-product stats)
     - 60-day Sales Efficiency     (overall per-rep KPIs)
     - Rep phone / home list
     - PM/EVE only: the PRIOR slot's ASSIGNED output(s) for drive chaining
2. Edit the CONFIG block: SLOT, the file paths, PRIOR_LOC_FILES, ROSTER
   (today's available reps), OUTPUT_FILE. Add any brand-new city to GEO.
3. Run:  python3 lead_engine.py
4. Open OUTPUT_FILE: assigned reps fill the Service Provider column; helper
   columns and a BENCH list (unused reps) are appended.

GOOD TO KNOW
------------
- Drive time is ESTIMATED (straight-line x1.3 / 35mph), not live routing.
  Any city missing from GEO is printed as a WARNING and can't be placed until
  you add its (lat, lon).
- SAP often pre-fills Service Provider with placeholder reps -> the engine
  IGNORES those and assigns fresh.
- To force a rep onto a lead, put "Send <rep>" (or "<rep> to run") in that
  lead's Notes -> the engine PINS that rep and re-optimizes everyone else.
  (This is also how to handle a late pickup lead.)
- Only Market == 'DC Metro' and non-BATHSYSTEM leads are assigned.
================================================================================
"""
import pandas as pd, numpy as np, re, math
from scipy.optimize import linear_sum_assignment
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================== CONFIG ==============================
# Edit this block each shift (or just tell Claude the slot, files, and roster).

SLOT            = 'PM'                              # 'AM' | 'PM' | 'EVE'  (informational)
LEAD_FILE       = 'leads_this_slot.xlsx'            # the slot's SAP lead export
POWER_RANK_FILE = 'Power_ranking.xlsx'              # 2-week per-product stats
SALES_60_FILE   = 'Sales_Efficiency_60day.xlsx'     # 60-day overall KPIs per rep
HOME_LIST_FILE  = 'Rep_Phone_List.xlsx'             # rep home locations
OUTPUT_FILE     = '/mnt/user-data/outputs/LEADS_ASSIGNED.xlsx'

# Prior-slot ASSIGNED sheets for drive chaining (MOST RECENT LAST - it wins):
#   AM  -> []
#   PM  -> ['AM_ASSIGNED.xlsx']
#   EVE -> ['AM_ASSIGNED.xlsx', 'PM_ASSIGNED.xlsx']
PRIOR_LOC_FILES = []

# Today's available reps (first + last name; aliases handled below).
ROSTER = [
    'aj doroci', 'brian mullarkey', 'dorothy cosby', 'gyasi lester',
    'jonathan lakomyj', 'hunter willson', 'stephen bryan', 'titus smith',
]

# ---- Tuning knobs (rarely change) ----
K         = 12      # product-leads before a rep's 2-week number outweighs their 60-day avg (higher = lean on 60-day)
K2        = 8       # stabilizes each rep's 60-day NSLI toward the team average
VOL_ADD   = 10000   # $ added to the TOP-volume rep's score; 60-day volume as its own standalone factor
FLOOR     = 5000    # minimum Eff. NSLI to be eligible for a BIG lead
DRIVE_PEN = 40      # $ penalty per drive-minute (drive is the lightest factor)
CEIL      = 120     # hard maximum drive minutes (outer sanity bound)

MARKET_KEEP    = 'DC Metro'
CANCEL_REP     = 'james dunn'   # a lead pre-issued to this name = CANCELLED -> never assign
SKIP_STATUS    = {'open'}       # lead Status values that mean "leave unassigned"
SKIP_CAMPAIGN  = {'install consultant'}  # Campaign values handled by install consultants, not our reps
DROP_PROD      = {'BATHSYSTEM'}
BOTTOM_EXCLUDE = {'raymond hieronimus', 'stephanie poteet', 'david golladay',
                  'gary halford', 'neil waranch'}   # always-excluded low producers

# Name aliases: any spelling -> canonical name (applied to ROSTER + location files + home list)
ALIASES = {
    'al kowalski':'wojciech kowalski', 'aj doroci':'ajet doroci', 'ben sands':'benjamin sands',
    'ray wander':'raymond wander', 'frank hill':'franklin hill', 'mary beth heller':'mary-beth heller',
    'marybeth heller':'mary-beth heller', 'nick wittman':'nicholas wittman', 'jeff kaelin':'jeffrey kaelin',
    'jp feeney':'john-paul feeney', 'john paul feeney':'john-paul feeney', 'johnpaul feeney':'john-paul feeney',
    'josh brown':'joshua brown', 'chris mercer':'christopher mercer', 'sam ludwig':'samantha ludwig',
    'steve forss':'steven forss', 'jaems archy':'james archy', 'zach diffenderfer':'zachary diffenderfer',
    'zach diff':'zachary diffenderfer', 'jacob szczepanik':'jacob szczepanik',
}

# Home overrides: new reps not yet in the phone-list export -> their home city.
HOME_OVERRIDES = {
    'jacob szczepanik':'crofton',
}

# City geocodes (lat, lon). ADD any new city here, or that lead/rep can't be placed.
GEO = {
    'washington':(38.9072,-77.0369),'baltimore':(39.2904,-76.6122),'silver spring':(38.9907,-77.0261),
    'rockville':(39.0840,-77.1528),'ellicott city':(39.2673,-76.7983),'bethesda':(38.9847,-77.0947),
    'upper marlboro':(38.8157,-76.7497),'sykesville':(39.3737,-76.9686),'reisterstown':(39.4690,-76.8294),
    'bowie':(38.9427,-76.7302),'boyds':(39.2143,-77.3216),'olney':(39.1532,-77.0669),'odenton':(39.0840,-76.7000),
    'waldorf':(38.6246,-76.9300),'temple hills':(38.8126,-76.9400),'gaithersburg':(39.1434,-77.2014),
    'lanham':(38.9676,-76.8636),'frederick':(39.4143,-77.4105),'owings mills':(39.4196,-76.7802),
    'clinton':(38.7651,-76.8983),'germantown':(39.1732,-77.2700),'alexandria':(38.8048,-77.0469),
    'winchester':(39.1857,-78.1633),'fairfax station':(38.8043,-77.3203),'district heights':(38.8576,-76.8894),
    'fort washington':(38.7099,-77.0300),'university park':(38.9701,-76.9447),'manassas':(38.7509,-77.4753),
    'accokeek':(38.6671,-76.9886),'davidsonville':(38.9290,-76.6308),'deale':(38.7793,-76.5494),
    'westminster':(39.5754,-76.9958),'potomac':(39.0182,-77.2086),'derwood':(39.1287,-77.1539),
    'falls church':(38.8823,-77.1711),'millersville':(39.0626,-76.6299),'thurmont':(39.6237,-77.4111),
    'clifton':(38.7807,-77.3866),'columbia':(39.2037,-76.8610),'centreville':(38.8401,-77.4291),
    'leesburg':(39.1157,-77.5636),'great mills':(38.2604,-76.4969),'gwynn oak':(39.3287,-76.7305),
    'hollywood':(38.3393,-76.5566),'upper falls':(39.4015,-76.4133),'vienna':(38.9012,-77.2653),
    'chevy chase':(38.9682,-77.0728),'chevrolet':(38.9682,-77.0728),
    'severna park':(39.0743,-76.5500),'herndon':(38.9696,-77.3861),'owings':(38.6900,-76.6100),
    'bel air':(39.5359,-76.3500),
    'annandale':(38.8304,-77.1964),'college park':(38.9807,-76.9369),'elkridge':(39.2126,-76.7136),
    'fulton':(39.1526,-76.9230),'gainesville':(38.7959,-77.6147),'king george':(38.2682,-77.1839),
    'lexington park':(38.2668,-76.4527),'woodbridge':(38.6582,-77.2497),
    'brunswick':(39.3140,-77.6249),'finksburg':(39.4940,-76.8800),'great falls':(38.9979,-77.2880),
    'hagerstown':(39.6418,-77.7200),'windsor mill':(39.3334,-76.7800),
    'brandywine':(38.6976,-76.8483),'ijamsville':(39.3132,-77.3589),'lake frederick':(39.0501,-78.1300),
    'mount airy':(39.3762,-77.1547),'suitland':(38.8487,-76.9197),
    'catonsville':(39.2720,-76.7319),'clarksville':(39.2037,-76.9469),'keedysville':(39.4862,-77.6961),
    'locust grove':(38.3457,-77.7600),'alexandria city':(38.8048,-77.0469),'catlett':(38.6543,-77.6386),
    'chantilly':(38.8943,-77.4311),'glen burnie':(39.1626,-76.6247),'haymarket':(38.8124,-77.6361),
    'stafford':(38.4221,-77.4083),
    'halethorpe':(39.2293,-76.6705),'lutherville timonium':(39.4243,-76.6190),'timonium':(39.4385,-76.6094),
    'sterling':(39.0062,-77.4286),
    'ashton':(39.1532,-77.0030),'brookeville':(39.1840,-77.0586),'clarksburg':(39.2380,-77.2786),
    'graysonville':(38.9460,-76.2107),
    'damascus':(39.2887,-77.2036),'mc lean':(38.9339,-77.1773),'mechanicsville':(38.4438,-76.7383),
    'montgomery village':(39.1768,-77.1953),'rosedale':(39.3243,-76.5097),'sparrows point':(39.2230,-76.4358),
    'cheverly':(38.9290,-76.9158),'saint leonard':(38.4757,-76.5125),'seat pleasant':(38.8968,-76.9047),
    'warrenton':(38.7135,-77.7950),'crofton':(39.0007,-76.6850),
    'hanover':(39.1888,-76.7236),'mclean':(38.9339,-77.1773),'oakton':(38.8810,-77.3000),
    'parkton':(39.6204,-76.6597),'takoma park':(38.9779,-77.0075),
    'burke':(38.7934,-77.2717),'essex':(39.3093,-76.4750),'gambrills':(39.0857,-76.6622),
    'new windsor':(39.5417,-77.1067),'queenstown':(38.9982,-76.1597),'woodsboro':(39.5337,-77.3175),
    'forest hill':(39.5840,-76.3950),'severn':(39.1372,-76.6983),'broadlands':(39.0157,-77.5347),
    'aldie':(38.9719,-77.6403),'havre de grace':(39.5495,-76.0902),'lovettsville':(39.2735,-77.6386),
    'stephenson':(39.2196,-78.0000),
    'capitol heights':(38.8851,-76.9136),'gambrels':(39.0857,-76.6622),'reston':(38.9586,-77.3570),
    'leonardtown':(38.2904,-76.6355),'baldwin':(39.5093,-76.4919),'white plains':(38.5893,-76.9789),
    'abingdon':(39.4668,-76.2986),'brambleton':(38.9821,-77.5389),'hampstead':(39.6046,-76.8497),
    'huntingtown':(38.6193,-76.6383),'nottingham':(39.3743,-76.4358),
    'bealeton':(38.5738,-77.7708),'edgewood':(39.4187,-76.2944),'round hill':(39.1362,-77.7597),
    'boonsboro':(39.5074,-77.6522),'bristow':(38.7573,-77.5397),'front royal':(38.9187,-78.1944),
    'aberdeen':(39.5096,-76.1641),'dundalk':(39.2507,-76.5205),'linthicum heights':(39.2054,-76.6525),
    'trappe':(38.6579,-76.0577),
    'nokesville':(38.6987,-77.5786),'belcamp':(39.4726,-76.2436),'churchville':(39.5640,-76.2483),
    'la plata':(38.5293,-76.9753),'manchester':(39.6612,-76.8850),'union bridge':(39.5690,-77.1772),
    'edgewater':(38.9568,-76.5494),'grasonville':(38.9460,-76.2107),'kensington':(39.0259,-77.0764),
    'north potomac':(39.0954,-77.2480),'pikesville':(39.3743,-76.7252),'randallstown':(39.3673,-76.7950),
    'ashburn':(39.0437,-77.4875),'cockeysville':(39.4812,-76.6427),'hillsboro':(39.2065,-77.7252),
    'hughesville':(38.5318,-76.7886),'jeffersonton':(38.6385,-77.8800),'lorton':(38.7043,-77.2278),
    # rep home cities
    'pasadena':(39.1071,-76.5694),'annapolis':(38.9784,-76.4922),'woodbine':(39.3362,-77.0639),
    'riva':(38.9568,-76.5808),'fairfax':(38.8462,-77.3064),'laurel':(39.0993,-76.8483),
    'beltsville':(39.0345,-76.9075),'middle river':(39.3343,-76.4391),'greenbelt':(38.9959,-76.8755),
    'springfield':(38.7893,-77.1872),'crownsville':(39.0285,-76.6041),'stevensville':(38.9818,-76.3314),
    'hyattsville':(38.9559,-76.9456),'chesapeake beach':(38.6857,-76.5336),'parkville':(39.3784,-76.5394),
    'arnold':(39.0326,-76.5030),'oxon hill':(38.8013,-76.9869),'burtonsville':(39.1116,-76.9325),
    'arlington':(38.8799,-77.1068),'chester':(37.3549,-77.4413),
}
# ============================ END CONFIG ============================


def canon(n):
    n = str(n).strip().lower()
    return ALIASES.get(n, n)

# ---- 2-week per-product stats (power ranking) ----
_raw = pd.read_excel(POWER_RANK_FILE, sheet_name='2 Weeks-Overall', header=None)
PS = {'Windows':19, 'Doors':35, 'Siding':51, 'Gutters':67, 'Roofing':83}; TOT = 3
def _gv(r, c):
    v = _raw.iloc[r, c]
    return float(v) if pd.notna(v) and not isinstance(v, str) else 0.0
reps = {}
for r in range(5, len(_raw)):
    fn = _raw.iloc[r, 0]
    if pd.isna(fn): continue
    nm = f"{str(fn).strip()} {str(_raw.iloc[r,1]).strip()}".lower()
    pr = {p: {'leads':_gv(r,c), 'nsli':_gv(r,c+15), 'close':_gv(r,c+12)} for p, c in PS.items()}
    pr['_tot'] = {'leads':_gv(r,TOT), 'nsli':_gv(r,TOT+15), 'close':_gv(r,TOT+12)}
    reps[nm] = pr

# ---- rep home locations ----
_hf = pd.read_excel(HOME_LIST_FILE, sheet_name='Sheet1', header=0)
_hf = _hf[_hf['Sales Representatives DMV'].notna()]
homes_map = {}
for _, _r in _hf.iterrows():
    _nm = canon(_r['Sales Representatives DMV'])
    _loc = str(_r['Location']).split(',')[0].strip().lower().replace(' bch', ' beach')
    if _loc and _loc not in ('location', 'nan'):
        homes_map[_nm] = _loc
for _k, _v in HOME_OVERRIDES.items():        # new reps not yet in the phone list
    homes_map[canon(_k)] = _v.strip().lower()

# ---- 60-day overall KPIs per rep ----
_s60 = pd.read_excel(SALES_60_FILE, sheet_name='Sales Efficiency All', header=1)
_s60 = _s60[_s60['Sales Representative Name'].notna()]
def _f(x): return float(x) if pd.notna(x) else 0.0
d60 = {}; _tn = 0.0; _ti = 0.0
for _, r in _s60.iterrows():
    nm = str(r['Sales Representative Name']).strip().lower()
    if nm == 'total': continue
    iss = _f(r['Issued']); net = _f(r['Closed $']) - _f(r['Rescinded Value']) - _f(r['Cancelled Value'])
    d60[nm] = {'nsli': net/iss if iss > 0 else 0.0, 'close': _f(r['Close %']), 'issued': iss, 'vol': net}
    _tn += net; _ti += iss
GLOBAL60 = _tn/_ti if _ti else 0.0
def get60(c): return d60.get(c)

# ---- available reps (with home), filtered to roster ----
ROSTER_C = [canon(x) for x in ROSTER]
_miss = [c for c in ROSTER_C if c not in reps or c not in homes_map]
if _miss: print('WARNING: roster reps with no ranking/home data (skipped):', _miss)
avail = [(c.title(), c, homes_map[c]) for c in ROSTER_C
         if c in reps and c in homes_map and c not in BOTTOM_EXCLUDE]

# ---- drive origins: prior-slot lead city (chained) else home ----
def _loadloc(fn):
    d = pd.read_excel(fn, header=0); m = {}
    for _, r in d.iterrows():
        city = str(r.get('City', '')).strip()
        if pd.notna(r.get('Service Provider')) and city and city.lower() != 'nan':
            k = str(r['Service Provider']).strip().lower()
            m[ALIASES.get(k, k)] = city
    return m
origin_map = {}
for fn in PRIOR_LOC_FILES:
    origin_map.update(_loadloc(fn))   # later (more recent) slot wins
avail = [(disp, c, origin_map.get(c, home)) for (disp, c, home) in avail]

# ---- team per-product prior (fallback for reps with no 60-day data) ----
prior = {}
for p in PS:
    tnum = sum(reps[c][p]['nsli']*reps[c][p]['leads'] for _, c, _ in avail)
    tot_l = sum(reps[c][p]['leads'] for _, c, _ in avail)
    prior[p] = tnum/tot_l if tot_l > 0 else 0.0
def rep_prior(c):
    g = get60(c)
    if not g: return None
    iss = g['issued']
    return (iss*g['nsli'] + K2*GLOBAL60) / (iss + K2)
def eff(c, p):
    d = reps[c][p]; l = d['leads']; rp = rep_prior(c)
    base = rp if rp is not None else prior[p]
    return (l*d['nsli'] + K*base) / (l + K)
def base_val(c, comps): return sum(eff(c, p) for p in comps) / len(comps)
def volnorm(c):
    g = get60(c); return max(0, g['vol'])/VOL_MAX if g else 0.0
def tiebreak(c):
    g = get60(c)
    if g: return 1e-3*g['close'] + 1e-7*g['nsli'] + 1e-10*g['issued']
    t = reps[c]['_tot']; return 1e-3*t['close'] + 1e-7*t['nsli'] + 1e-10*t['leads']

# ---- drive-time estimate ----
def _hav(a, b):
    R = 3958.8; la1, lo1 = map(math.radians, a); la2, lo2 = map(math.radians, b)
    dd = math.sin((la2-la1)/2)**2 + math.cos(la1)*math.cos(la2)*math.sin((lo2-lo1)/2)**2
    return 2*R*math.asin(math.sqrt(dd))
def drive(c1, c2):
    g1 = GEO.get(str(c1).strip().lower()); g2 = GEO.get(str(c2).strip().lower())
    return None if not g1 or not g2 else _hav(g1, g2)*1.3/35*60

# ---- note parsing: window / door counts + product components ----
W = {'one':1,'two':2,'three':3,'four':4,'five':5,'six':6,'seven':7,'eight':8,'nine':9,'ten':10,
     'eleven':11,'twelve':12,'thirteen':13,'fourteen':14,'fifteen':15,'twenty':20}
NUM = r'(\d{1,2}|one|two|three|four|five|six|seven|eight|nine|ten|eleven|twelve|thirteen|fourteen|fifteen|twenty)'
def _nn(t): t = t.strip().lower(); return int(t) if t.isdigit() else W.get(t, 0)
def windows(s):
    nums = []
    for a, b in re.findall(NUM + r'\s*(?:to|through|-|–)\s*' + NUM, s):
        if 'window' in s: nums += [_nn(a), _nn(b)]
    for m in re.findall(r'(?<!\d)' + NUM + r'(?:[\s,]+[\w,-]+){0,2}?[\s,]+windows?', s): nums.append(_nn(m))
    for m in re.findall(r'(?<!\d)(\d{1,2})\s*windows?', s): nums.append(int(m))
    for m in re.findall(r'(?<!\d)' + NUM + r'\s+win\b', s): nums.append(_nn(m))
    nums = [x for x in nums if 0 < x <= 40]
    base = max(nums) if nums else 0
    bb = 0
    for m in re.findall(r'(?<!\d)' + NUM + r'\s+(?:bow|bay)\s+windows?', s): bb = max(bb, _nn(m))
    if bb == 0 and re.search(r'(?:bow|bay)\s+window', s): bb = 1
    total = base + bb if (bb and re.search(r'(?:additional|more|other)\s+windows?', s)) else max(base, bb)
    return (total, bool(nums) or bb > 0)
def doors(s):
    s = re.sub(r'(\d)([a-z])', r'\1 \2', s)   # unglue '2storms' -> '2 storms'
    def cnt(kw, dflt=1):
        ns = []
        for m in re.finditer(r'(?<!\d)' + NUM + r'(?:[\s,]+[\w,-]+){0,2}?[\s,]+' + kw, s): ns.append(_nn(m.group(1)))
        for m in re.finditer(r'(\d{1,2})\s*' + kw, s): ns.append(int(m.group(1)))
        ns = [x for x in ns if 0 < x <= 20]
        if ns: return max(ns)
        return dflt if re.search(kw, s) else 0
    front = cnt(r'(?:fed\b|front\s+entry\s+doors?|front\s+doors?|entry\s+doors?)')
    for _m in re.findall(r'(?<!\d)' + NUM + r'\s+ed\b', s): front = max(front, _nn(_m))
    storm = cnt(r'storms?(?:\s+doors?)?\b')
    slide = cnt(r'(?:sgd\b|sliding\s+glass\s+doors?|sliding\s+patio\s+doors?|patio\s+doors?|sliding\s+doors?)')
    generic = cnt(r'doors?\b', 0)
    t = set()
    if front > 0: t.add('front')
    if storm > 0: t.add('storm')
    if slide > 0: t.add('sliding')
    d = max(generic, front + storm + slide)
    return d, t, (d > 0 or bool(t))
def bare(s):
    c = re.findall(r'(?<![\d/@:.])\b(\d{1,2})\b(?!\s*(?:pm|am|st|nd|rd|th)\b)(?![\d/:])', s)
    c = [int(x) for x in c if 0 < int(x) <= 40]
    return max(c) if c else 0

# ---- product -> components used for SCORING (bath/extra components shown only as tags) ----
PRODC = {
    'WINDOWS':['Windows'], 'DOORS':['Doors'], 'ROOFING':['Roofing'], 'SIDING':['Siding'], 'GUTTERS':['Gutters'],
    'COMBO_WIND_DOOR':['Windows','Doors'], 'COMBO_WIN_DOR_ROOF':['Windows','Doors','Roofing'],
    'COMBO_S_W_G_D':['Siding','Windows','Gutters','Doors'], 'COMBO_WIN_DOOR_GUT':['Windows','Doors','Gutters'],
    'COMBO_SID_GUT_ROOF':['Siding','Gutters','Roofing'], 'COMBO_BATH_WIN_DOR':['Windows','Doors'],
    'COMBO_WIN_SID_ROOF':['Windows','Siding','Roofing'], 'COMBO_WIN_ROOFING':['Windows','Roofing'],
    'COMBO_SIDING_WIND':['Windows'], 'COMBO_BATH_W_D_G_S':['Windows','Doors'],
    'COMBO_SID_WIN_DOOR':['Windows','Doors'], 'COMBO_WIND_GUTTER':['Windows'],
    'COMBO_SIDING_GUTT':['Siding','Gutters'],
    'COMBO_DOOR_ROOFING':['Doors','Roofing'],
    'COMBO_S_W_G_D_R':['Siding','Windows','Gutters','Doors','Roofing'],
    'COMBO_SID_DOOR_GUT':['Siding','Doors','Gutters'],
    'COMBO_SIDING_DOOR':['Siding','Doors'],
    'COMBO_GUT_ROOFING':['Gutters','Roofing'],
    'COMBO_SID_ROOFING':['Siding','Roofing'],
}

# ---- opportunity tier ----
def tier(prod, w, wk, d, dt):
    c = PRODC.get(prod, [prod]); nc = len(c); pu = str(prod).upper()
    tk = [t for t in pu.split('_') if t not in ('COMBO', 'BATH')]
    sid  = 'Siding'  in c or 'SIDING'  in pu or 'S' in tk or 'SID' in tk
    roof = 'Roofing' in c or 'ROOF'    in pu or 'R' in tk
    gut  = 'Gutters' in c or 'GUTTER'  in pu or 'G' in tk or 'GUT' in tk
    nc = max(nc, len(tk))
    wincombo = pu.startswith('COMBO') and ('WIN' in pu or 'W' in tk)
    if sid: return 'BIG'
    if nc >= 3: return 'BIG'
    if w >= 7: return 'BIG'
    if d >= 3: return 'BIG'
    if w >= 5 and d >= 2: return 'BIG'
    if roof: return 'MEDIUM'
    if gut and nc >= 2: return 'MEDIUM'
    if w == 6: return 'MEDIUM'
    if d == 2: return 'MEDIUM'
    if d == 1 and 'front' in dt: return 'MEDIUM'
    return 'MEDIUM' if wincombo else 'SMALL'   # any window-combo is at least MEDIUM

# ============================ ASSIGNMENT ============================
pmdf = pd.read_excel(LEAD_FILE, header=0)
_spcol = pmdf.columns.get_loc('Service Provider')
# capture cancellation / skip signals BEFORE clearing placeholder reps
_orig_sp = pmdf['Service Provider'].astype(str).str.strip().str.lower()
_status  = (pmdf['Status'].astype(str).str.strip().str.lower()
            if 'Status' in pmdf.columns else pd.Series(['']*len(pmdf)))
_campaign = (pmdf['Campaign'].astype(str).str.strip().str.lower()
             if 'Campaign' in pmdf.columns else pd.Series(['']*len(pmdf)))
pmdf['Service Provider'] = pd.Series([np.nan]*len(pmdf), dtype=object)   # ignore SAP-prefilled placeholders
drop_idx = [i for i in range(len(pmdf))
            if (str(pmdf.iloc[i]['Market']).strip() != MARKET_KEEP)
            or (pmdf.iloc[i]['Product'] in DROP_PROD)
            or (_orig_sp.iloc[i] == CANCEL_REP)        # James Dunn = cancelled
            or (_status.iloc[i] in SKIP_STATUS)        # Status 'Open' = leave unassigned
            or (_campaign.iloc[i] in SKIP_CAMPAIGN)]   # Install Consultant campaign = not our rep
_n_cancel = sum(1 for i in range(len(pmdf)) if _orig_sp.iloc[i] == CANCEL_REP)
_n_open   = sum(1 for i in range(len(pmdf)) if _status.iloc[i] in SKIP_STATUS)
_n_camp   = sum(1 for i in range(len(pmdf)) if _campaign.iloc[i] in SKIP_CAMPAIGN)
if _n_cancel or _n_open or _n_camp:
    print(f"SKIPPED: {_n_cancel} cancelled ({CANCEL_REP}), {_n_open} open-status, {_n_camp} install-consultant")

# NOTE-PIN: honor 'send/assign/give/have <rep>' or '<rep> to run' in a lead's Notes
_byname = {}
for _d, _c, _o in avail:
    for _p in (_c.split()[-1], _c.split()[0]): _byname.setdefault(_p, set()).add(_c)
note_pins = {}
for i in range(len(pmdf)):
    if i in drop_idx: continue
    _nt = str(pmdf.iloc[i]['Notes']).lower()
    _hits = re.findall(r'\b(?:send|assign|give|have|book)\s+(?:it\s+|this\s+|to\s+)*([a-z]+)', _nt) \
            + re.findall(r'\b([a-z]+)\s+to\s+run\b', _nt)
    for _tok in _hits:
        cn = _byname.get(_tok)
        if cn and len(cn) == 1:
            rep = next(iter(cn)); pmdf.iat[i, _spcol] = rep.title(); note_pins[i] = rep; break
if note_pins:
    print('NOTE-PINS:', {pmdf.iloc[i]['Customer Name']: r for i, r in note_pins.items()})

open_idx = [i for i in range(len(pmdf)) if pd.isna(pmdf.iloc[i]['Service Provider']) and i not in drop_idx]
pre_reps = {canon(pmdf.iloc[i]['Service Provider']) for i in note_pins}
avail = [a for a in avail if a[1] not in pre_reps]

# missing-geocode warning
_need = set(str(pmdf.iloc[i]['City']).strip().lower() for i in range(len(pmdf)) if i not in drop_idx)
_need |= set(str(o).strip().lower() for _, _, o in avail)
_missgeo = sorted(x for x in _need if x and x not in GEO)
if _missgeo: print('WARNING: no geocode for these cities (add to GEO):', _missgeo)

# parse every open + pinned lead
L = {}
for i in open_idx + list(note_pins):
    row = pmdf.iloc[i]; s = str(row['Notes']).lower().split('pre-cust')[0]
    w, wk = windows(s); d, dt, dk = doors(s); prod = row['Product']
    if 'WIN' in str(prod).upper() and w == 0:          # window-combo count fallback (incl ranges like 6-8)
        rng = re.findall(r'(\d+)\s*[-–]\s*(\d+)', s)
        if rng: w, wk = max(int(b) for a, b in rng), True
        else:
            bn = bare(s)
            if bn: w, wk = bn, True
    if prod in ('DOORS',) and d == 0 and not dt:
        bn = bare(s)
        if bn: d, dk = bn, True
    comps = PRODC.get(prod, ['Windows']); pu = str(prod).upper()
    tk = [t for t in pu.split('_') if t not in ('COMBO', 'BATH')]
    L[i] = {'comps':comps, 'w':w, 'd':d, 'tier':tier(prod, w, wk, d, dt), 'city':row['City'], 'prod':prod,
            'sid': ('Siding'  in comps or re.search(r'\bsiding', s)  or 'SIDING'  in pu or 'S' in tk or 'SID' in tk),
            'roof':('Roofing' in comps or re.search(r'\broof', s)    or 'ROOF'    in pu or 'R' in tk),
            'gut': ('Gutters' in comps or re.search(r'\bgutter', s)  or 'GUTTER'  in pu or 'G' in tk or 'GUT' in tk)}

VOL_MAX = max([get60(c)['vol'] for _, c, _ in avail if get60(c) and get60(c)['vol'] > 0], default=1.0)

# tier-ordered constrained optimization (BIG first, then MEDIUM, then SMALL)
pool = list(range(len(avail))); res = {}
for T in ['BIG', 'MEDIUM', 'SMALL']:
    leads = [i for i in open_idx if L[i]['tier'] == T]
    if not leads or not pool:
        for i in leads: res[i] = None
        continue
    M = np.full((len(leads), len(pool)), -1e9); DRm = np.full((len(leads), len(pool)), np.nan)
    for a, i in enumerate(leads):
        for b, ri in enumerate(pool):
            rr, c, org = avail[ri]; dm = drive(org, L[i]['city']); DRm[a, b] = dm if dm is not None else np.nan
            if dm is None or dm > CEIL: continue
            bv = base_val(c, L[i]['comps'])
            if T == 'BIG' and bv < FLOOR: continue
            M[a, b] = bv + VOL_ADD*volnorm(c) - DRIVE_PEN*dm + tiebreak(c)
    rows, colsx = linear_sum_assignment(-M); used = []
    for a, b in zip(rows, colsx):
        i = leads[a]; ri = pool[b]
        if M[a, b] > -1e8:
            rr, c, org = avail[ri]; res[i] = (ri, DRm[a, b], base_val(c, L[i]['comps'])); used.append(ri)
        else: res[i] = None
    for i in leads:
        if i not in res: res[i] = None
    pool = [ri for ri in pool if ri not in used]

def feas_all(i, floor=False):
    out = []
    for ri in range(len(avail)):
        rr, c, org = avail[ri]; dm = drive(org, L[i]['city'])
        if dm is None or dm > CEIL: continue
        if floor and base_val(c, L[i]['comps']) < FLOOR: continue
        out.append(rr.title())
    return out

# ============================ WRITE OUTPUT ============================
wb = load_workbook(LEAD_FILE); ws = wb['SAPUI5 Export']
hdr = [c.value for c in ws[1]]; sp = hdr.index('Service Provider') + 1; base = len(hdr)
for r in range(2, ws.max_row + 1): ws.cell(r, sp).value = None   # clear SAP placeholders
cols = ['Tier', 'Opportunity', 'Rep Home', 'Est Drive (min)', 'Eff. NSLI ($)', 'Assignment Flag']
hf = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF', name='Arial')
for j, h in enumerate(cols):
    c = ws.cell(1, base+1+j, h); c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal='center', wrap_text=True)
for col in range(1, base+1):
    cc = ws.cell(1, col); cc.fill = hf; cc.font = hfont; cc.alignment = Alignment(horizontal='center', wrap_text=True)
ws.row_dimensions[1].height = 30
warn = PatternFill('solid', fgColor='FFF2CC'); bad = PatternFill('solid', fgColor='F8CBAD')
biggrn = PatternFill('solid', fgColor='E2EFDA'); pinfill = PatternFill('solid', fgColor='D9E1F2')

def _opp(d):
    parts = [f"{d['w']}W" if d['w'] else '', f"{d['d']}D" if d['d'] else '']
    if d['sid']: parts.append('1 siding')
    if d['roof']: parts.append('1 roof')
    if d['gut']: parts.append('1 Gutt')
    return ' '.join(x for x in parts if x) or '—'

for i in open_idx:
    er = i+2; d = L[i]
    ws.cell(er, base+1, d['tier']); ws.cell(er, base+2, _opp(d)); flags = []
    if res[i]:
        ri, dm, bv = res[i]; rr, c, org = avail[ri]
        ws.cell(er, sp, rr.title()); ws.cell(er, base+3, org.title())
        ws.cell(er, base+4, round(dm)); ws.cell(er, base+5, round(bv))
        flags.append(f'from: {org.title()}')
        g = get60(c)
        if g: flags.append(f"60-day vol ${g['vol']/1000:.0f}K")
        if 'BATH' in str(d['prod']).upper(): flags.append('contains bath component — scored on windows/doors')
        if dm > 75: flags.append(f'verify drive ~{round(dm)}m')
        thin = all(reps[c][p]['leads'] < 5 for p in d['comps'])
        if thin: flags.append('thin 2-wk product volume — leaning on 60-day overall')
        iss = int(g['issued']) if g else 0
        if d['tier'] == 'BIG' and iss < 10: flags.append(f'limited 60-day volume ({iss} issued) — verify for BIG lead')
        rf = warn if (dm > 75 or thin) else (biggrn if d['tier'] == 'BIG' else None)
        if rf:
            for k in range(1, base+7): ws.cell(er, k).fill = rf
    else:
        fa = feas_all(i, floor=(d['tier'] == 'BIG'))
        if d['tier'] == 'BIG' and not fa and feas_all(i):
            flags.append(f'UNFILLED — reachable reps all below ${FLOOR:,} BIG floor')
        else:
            flags.append('UNCOVERED — only reachable: ' + ', '.join(fa) + ' (taken)' if fa
                         else 'UNCOVERED — no rep within %dm' % CEIL)
        for k in range(1, base+7): ws.cell(er, k).fill = bad
    ws.cell(er, base+6, '; '.join(flags))

for i, c in note_pins.items():        # pinned (note / late-lead) rows
    er = i+2; d = L[i]
    org = origin_map.get(c, homes_map.get(c, '')); dm = drive(org, d['city']); bv = base_val(c, d['comps'])
    ws.cell(er, sp, c.title()); ws.cell(er, base+1, d['tier']); ws.cell(er, base+2, _opp(d))
    if org: ws.cell(er, base+3, org.title())
    if dm is not None: ws.cell(er, base+4, round(dm))
    ws.cell(er, base+5, round(bv)); ws.cell(er, base+6, f'PINNED per note -> {c.title()}')
    for k in range(1, base+7): ws.cell(er, k).fill = pinfill

for i in sorted(drop_idx, reverse=True): ws.delete_rows(i+2, 1)
ws.auto_filter.ref = f"A1:{ws.cell(1, base+len(cols)).column_letter}{ws.max_row}"
ws.freeze_panes = 'A2'

# ---- BENCH: roster reps with no lead today (by 60-day volume) ----
cn = hdr.index('Customer Name') + 1
used_ri = {res[i][0] for i in res if res[i]}
def _vol(c): g = get60(c); return g['vol'] if g else 0
bench = sorted([avail[ri] for ri in range(len(avail)) if ri not in used_ri], key=lambda a: -_vol(a[1]))
r0 = ws.max_row + 2
hc = ws.cell(r0, cn, f'ON THE BENCH — {len(bench)} reps, no lead today (by 60-day volume)')
hc.font = Font(bold=True, color='FFFFFF', name='Arial'); hc.fill = hf
for k in [cn, sp, base+3, base+6]: ws.cell(r0, k).fill = hf
ws.cell(r0, sp, 'Rep').font = Font(bold=True, color='FFFFFF', name='Arial')
ws.cell(r0, base+3, 'Home').font = Font(bold=True, color='FFFFFF', name='Arial')
benchfill = PatternFill('solid', fgColor='F2F2F2')
for n, (rr, c, org) in enumerate(bench):
    rr_ = r0+1+n
    ws.cell(rr_, sp, rr.title()); ws.cell(rr_, base+3, org.title())
    v = _vol(c); ws.cell(rr_, base+6, f"60-day vol ${v/1000:.0f}K" if v else "")
    for k in [sp, base+3, base+6]: ws.cell(rr_, k).fill = benchfill

# ---- borders on every used cell + auto-fit column widths ----
thin = Side(style='thin', color='808080'); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
lastcol = base + len(cols)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=lastcol):
    for cc in row: cc.border = bd
wrapcols = {hdr.index('Notes')+1 if 'Notes' in hdr else 0, lastcol}
for col in range(1, lastcol+1):
    mx = 0
    for rr_ in range(1, ws.max_row+1):
        v = ws.cell(rr_, col).value
        if v is not None: mx = max(mx, max((len(x) for x in str(v).split(chr(10))), default=0))
    cap = 55 if col in wrapcols else 34
    ws.column_dimensions[ws.cell(1, col).column_letter].width = min(max(mx+2, 9), cap)
    if col in wrapcols:
        for rr_ in range(2, ws.max_row+1):
            cell = ws.cell(rr_, col)
            if cell.value is not None: cell.alignment = Alignment(wrap_text=True, vertical='top')

wb.save(OUTPUT_FILE)
print(f"SLOT={SLOT}  assigned={sum(1 for i in res if res[i])}  pinned={len(note_pins)}  bench={len(bench)}")
print('saved ->', OUTPUT_FILE)
